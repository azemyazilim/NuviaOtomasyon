from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db.models import Sum, Count, F
from datetime import date, datetime, timedelta
from openpyxl import Workbook
from reportlab.pdfgen import canvas
from satis.models import Satis, SatisDetay
from urun.models import Urun
from musteri.models import Musteri


@login_required
def gunluk_satis(request):
    """Günlük satış raporu view'ı"""
    bugun = date.today()
    tarih = request.GET.get('tarih', bugun.strftime('%Y-%m-%d'))
    
    try:
        secili_tarih = datetime.strptime(tarih, '%Y-%m-%d').date()
    except ValueError:
        secili_tarih = bugun
    
    # Günlük satışlar
    satislar = Satis.objects.filter(
        satis_tarihi__date=secili_tarih,
        durum='tamamlandi'
    )
    
    # İstatistikler
    toplam_satis = satislar.aggregate(
        toplam=Sum('toplam_tutar'),
        adet=Count('id')
    )
    
    context = {
        'satislar': satislar,
        'secili_tarih': secili_tarih,
        'toplam_satis': toplam_satis,
    }
    return render(request, 'rapor/gunluk_satis.html', context)


@login_required
def stok_raporu(request):
    """Stok raporu view'ı"""
    urunler = Urun.objects.filter(aktif=True).order_by('kategori__ust_kategori__ad', 'kategori__ad', 'ad')
    
    # Filtreler
    durum = request.GET.get('durum')
    if durum == 'tukendi':
        urunler = urunler.filter(stok_miktari=0)
    elif durum == 'kritik':
        urunler = urunler.filter(stok_miktari__gt=0, stok_miktari__lte=F('minimum_stok'))
    
    context = {
        'urunler': urunler,
        'durum': durum,
    }
    return render(request, 'rapor/stok_raporu.html', context)


@login_required
def cok_satan_urunler(request):
    """En çok satan ürünler raporu view'ı"""
    # Tarih aralığı
    baslangic = request.GET.get('baslangic')
    bitis = request.GET.get('bitis')
    
    if not baslangic:
        # Varsayılan: Bu ay
        baslangic = date.today().replace(day=1)
    else:
        baslangic = datetime.strptime(baslangic, '%Y-%m-%d').date()
    
    if not bitis:
        bitis = date.today()
    else:
        bitis = datetime.strptime(bitis, '%Y-%m-%d').date()
    
    # En çok satan ürünler
    cok_satanlar = SatisDetay.objects.filter(
        satis__satis_tarihi__date__range=[baslangic, bitis],
        satis__durum='tamamlandi'
    ).values('urun').annotate(
        toplam_miktar=Sum('miktar'),
        toplam_ciro=Sum('toplam_fiyat')
    ).order_by('-toplam_miktar')[:20]
    
    # Ürün bilgilerini ekle
    for item in cok_satanlar:
        item['urun_obj'] = Urun.objects.get(pk=item['urun'])
    
    context = {
        'cok_satanlar': cok_satanlar,
        'baslangic': baslangic,
        'bitis': bitis,
    }
    return render(request, 'rapor/cok_satan_urunler.html', context)


@login_required
def kar_zarar(request):
    """Kâr/Zarar analizi view'ı"""
    # Tarih aralığı
    baslangic = request.GET.get('baslangic')
    bitis = request.GET.get('bitis')
    
    if not baslangic:
        baslangic = date.today().replace(day=1)
    else:
        baslangic = datetime.strptime(baslangic, '%Y-%m-%d').date()
    
    if not bitis:
        bitis = date.today()
    else:
        bitis = datetime.strptime(bitis, '%Y-%m-%d').date()
    
    # Satış detayları
    satis_detaylari = SatisDetay.objects.filter(
        satis__satis_tarihi__date__range=[baslangic, bitis],
        satis__durum='tamamlandi'
    )
    
    # Kâr/Zarar hesaplama
    toplam_ciro = 0
    toplam_maliyet = 0
    
    for detay in satis_detaylari:
        toplam_ciro += detay.toplam_fiyat
        toplam_maliyet += (detay.urun.alis_fiyati * detay.miktar)
    
    toplam_kar = toplam_ciro - toplam_maliyet
    kar_marji = (toplam_kar / toplam_ciro * 100) if toplam_ciro > 0 else 0
    
    context = {
        'toplam_ciro': toplam_ciro,
        'toplam_maliyet': toplam_maliyet,
        'toplam_kar': toplam_kar,
        'kar_marji': kar_marji,
        'baslangic': baslangic,
        'bitis': bitis,
    }
    return render(request, 'rapor/kar_zarar.html', context)


@login_required
def musteri_raporu(request):
    """Müşteri raporu view'ı"""
    musteriler = Musteri.objects.filter(aktif=True)
    
    # Müşteri satış istatistikleri
    musteri_stats = []
    for musteri in musteriler:
        satislar = Satis.objects.filter(musteri=musteri, durum='tamamlandi')
        toplam_satis = satislar.aggregate(toplam=Sum('toplam_tutar'))['toplam'] or 0
        satis_adedi = satislar.count()
        
        if satis_adedi > 0:
            musteri_stats.append({
                'musteri': musteri,
                'toplam_satis': toplam_satis,
                'satis_adedi': satis_adedi,
                'ortalama_satis': toplam_satis / satis_adedi
            })
    
    # En çok alışveriş yapan müşteriler
    musteri_stats.sort(key=lambda x: x['toplam_satis'], reverse=True)
    
    context = {
        'musteri_stats': musteri_stats[:20],
    }
    return render(request, 'rapor/musteri_raporu.html', context)


# Excel Export Views
@login_required
def gunluk_satis_excel(request):
    """Günlük satış Excel export"""
    tarih = request.GET.get('tarih', date.today().strftime('%Y-%m-%d'))
    secili_tarih = datetime.strptime(tarih, '%Y-%m-%d').date()
    
    satislar = Satis.objects.filter(
        satis_tarihi__date=secili_tarih,
        durum='tamamlandi'
    )
    
    # Excel dosyası oluştur
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = f"Günlük Satış - {secili_tarih}"
    
    # Başlıklar
    headers = ['Satış No', 'Müşteri', 'Toplam Tutar', 'Ödeme Tipi', 'Tarih']
    for col, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col, value=header)
    
    # Veriler
    for row, satis in enumerate(satislar, 2):
        worksheet.cell(row=row, column=1, value=satis.satis_no)
        worksheet.cell(row=row, column=2, value=satis.musteri.tam_ad if satis.musteri else 'Bilinmeyen')
        worksheet.cell(row=row, column=3, value=float(satis.toplam_tutar))
        worksheet.cell(row=row, column=4, value=satis.get_odeme_tipi_display())
        worksheet.cell(row=row, column=5, value=satis.satis_tarihi.strftime('%d.%m.%Y %H:%M'))
    
    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="gunluk_satis_{secili_tarih}.xlsx"'
    workbook.save(response)
    return response


@login_required
def gunluk_satis_pdf(request):
    """Günlük satış PDF export"""
    # PDF oluşturma kodu buraya gelecek
    pass


@login_required
def stok_excel(request):
    """Stok raporu Excel export"""
    urunler = Urun.objects.filter(aktif=True)
    
    # Excel dosyası oluştur
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Stok Raporu"
    
    # Başlıklar
    headers = ['Barkod', 'Ürün Adı', 'Kategori', 'Stok Miktarı', 'Minimum Stok', 'Alış Fiyatı', 'Satış Fiyatı']
    for col, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col, value=header)
    
    # Veriler
    for row, urun in enumerate(urunler, 2):
        worksheet.cell(row=row, column=1, value=urun.barkod)
        worksheet.cell(row=row, column=2, value=urun.ad)
        worksheet.cell(row=row, column=3, value=str(urun.kategori))
        worksheet.cell(row=row, column=4, value=urun.stok_miktari)
        worksheet.cell(row=row, column=5, value=urun.minimum_stok)
        worksheet.cell(row=row, column=6, value=float(urun.alis_fiyati))
        worksheet.cell(row=row, column=7, value=float(urun.satis_fiyati))
    
    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="stok_raporu.xlsx"'
    workbook.save(response)
    return response


@login_required
def stok_pdf(request):
    """Stok raporu PDF export"""
    # PDF oluşturma kodu buraya gelecek
    pass


@login_required
def kar_zarar_excel(request):
    """Kâr/Zarar Excel export"""
    # Excel oluşturma kodu buraya gelecek
    pass


@login_required
def kar_zarar_pdf(request):
    """Kâr/Zarar PDF export"""
    # PDF oluşturma kodu buraya gelecek
    pass
